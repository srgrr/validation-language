from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from validation_language_mockup.evaluator import compile_rule, row_violation_mask
from validation_language_mockup.excel_export import export_validation_excel
from validation_language_mockup.parser import parse_avl

_VIOLATION_FILL = "FFC7CE"
_OK_FILL = "FFFFFF"


def test_row_violation_mask_when_not_matched() -> None:
    df = pl.DataFrame({"Origin": ["Madrid", "Barcelona"], "Destination": ["X", "Y"]})
    rule = parse_avl(
        """\
WHEN
    COL(Origin) = "Barcelona"
THEN
    COL(Destination) = "Z"
GROUP BY
    "Origin"
"""
    )
    compiled = compile_rule(rule)
    assert row_violation_mask(df, compiled) == [False, True]


def test_export_validation_excel_colors_rows(tmp_path: Path) -> None:
    df = pl.DataFrame(
        {
            "Origin": ["Madrid", "Barcelona"],
            "Destination": ["Madrid", "Barcelona"],
        }
    )
    rule = parse_avl(
        """\
WHEN
    COL(Origin) = "Barcelona"
THEN
    COL(Destination) != "Barcelona"
GROUP BY
    "Origin"
"""
    )
    compiled = compile_rule(rule)
    out = tmp_path / "out.xlsx"

    count = export_validation_excel(df, compiled, out)

    assert count == 1
    assert out.is_file()

    ws = load_workbook(out).active
    assert ws.cell(1, 1).value == "Origin"
    assert ws.cell(1, 1).fill.fgColor.rgb in (None, "00000000", "FFFFFFFF")

    ok_fill = ws.cell(2, 1).fill.fgColor.rgb
    bad_fill = ws.cell(3, 1).fill.fgColor.rgb
    assert ok_fill.endswith(_OK_FILL) or ok_fill in ("FFFFFFFF", "00FFFFFF")
    assert bad_fill.endswith(_VIOLATION_FILL)
