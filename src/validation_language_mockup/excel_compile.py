"""Compile AVL rules to Excel formulas for conditional formatting."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.utils import get_column_letter

from validation_language_mockup.ast import (
    AllEqualExpr,
    AllExpr,
    AnyExpr,
    BoolAnd,
    BoolExpr,
    BoolFalse,
    BoolNot,
    BoolOr,
    BoolTrue,
    ColExpr,
    CompareExpr,
    CompareOp,
    Expr,
    IntLiteral,
    IsNullExpr,
    Rule,
    StringLiteral,
)

_EXCEL_COMPARE: dict[CompareOp, str] = {
    "<": "<",
    "<=": "<=",
    ">": ">",
    ">=": ">=",
    "=": "=",
    "!=": "<>",
}


@dataclass(frozen=True)
class ExcelCompileContext:
    """Column letters and template row for Excel conditional-formatting formulas."""

    columns: dict[str, str]
    row: int = 2

    @classmethod
    def from_column_names(cls, names: list[str], *, row: int = 2) -> ExcelCompileContext:
        letters = {
            name: get_column_letter(index + 1) for index, name in enumerate(names)
        }
        return cls(columns=letters, row=row)

    def cell(self, name: str) -> str:
        try:
            letter = self.columns[name]
        except KeyError as exc:
            msg = f"column {name!r} not found in CSV"
            raise ValueError(msg) from exc
        return f"{letter}{self.row}"

    def col_range(self, name: str) -> str:
        try:
            letter = self.columns[name]
        except KeyError as exc:
            msg = f"column {name!r} not found in CSV"
            raise ValueError(msg) from exc
        return f"${letter}:${letter}"


def _excel_string(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def compile_violation_formula(rule: Rule, column_names: list[str], *, row: int = 2) -> str:
    """Excel formula that is TRUE when a row is a validation violation."""
    ctx = ExcelCompileContext.from_column_names(column_names, row=row)
    group_by = list(rule.group_by)
    when = compile_excel_bool(rule.when, ctx, group_by=group_by)
    then = compile_excel_bool(rule.then, ctx, group_by=group_by)
    return f"AND({when},NOT({then}))"


def compile_excel_bool(
    expr: BoolExpr,
    ctx: ExcelCompileContext,
    *,
    group_by: list[str],
) -> str:
    if isinstance(expr, BoolTrue):
        return "TRUE"
    if isinstance(expr, BoolFalse):
        return "FALSE"
    if isinstance(expr, BoolNot):
        return f"NOT({compile_excel_bool(expr.operand, ctx, group_by=group_by)})"
    if isinstance(expr, BoolAnd):
        left = compile_excel_bool(expr.left, ctx, group_by=group_by)
        right = compile_excel_bool(expr.right, ctx, group_by=group_by)
        return f"AND({left},{right})"
    if isinstance(expr, BoolOr):
        left = compile_excel_bool(expr.left, ctx, group_by=group_by)
        right = compile_excel_bool(expr.right, ctx, group_by=group_by)
        return f"OR({left},{right})"
    if isinstance(expr, CompareExpr):
        left = compile_excel_value(expr.left, ctx)
        right = compile_excel_value(expr.right, ctx)
        return f"{left}{_EXCEL_COMPARE[expr.op]}{right}"
    if isinstance(expr, AllEqualExpr):
        return compile_excel_all_equal(expr, ctx, group_by=group_by)
    if isinstance(expr, AnyExpr):
        if not expr.operands:
            return "FALSE"
        parts = [
            compile_excel_bool(operand, ctx, group_by=group_by) for operand in expr.operands
        ]
        return f"OR({','.join(parts)})"
    if isinstance(expr, AllExpr):
        if not expr.operands:
            return "TRUE"
        parts = [
            compile_excel_bool(operand, ctx, group_by=group_by) for operand in expr.operands
        ]
        return f"AND({','.join(parts)})"
    if isinstance(expr, IsNullExpr):
        blank = f"ISBLANK({ctx.cell(expr.col.name)})"
        return f"NOT({blank})" if expr.negated else blank
    if isinstance(expr, ColExpr):
        return f"NOT(ISBLANK({ctx.cell(expr.name)}))"
    if isinstance(expr, IntLiteral):
        return str(expr.value)
    if isinstance(expr, StringLiteral):
        return _excel_string(expr.value)
    msg = f"unsupported boolean expression: {type(expr).__name__}"
    raise TypeError(msg)


def compile_excel_all_equal(
    expr: AllEqualExpr,
    ctx: ExcelCompileContext,
    *,
    group_by: list[str],
) -> str:
    if not group_by:
        msg = "ALL_EQUAL requires at least one GROUP BY column"
        raise ValueError(msg)
    key_criteria: list[str] = []
    for key in group_by:
        key_criteria.extend([ctx.col_range(key), ctx.cell(key)])
    value_col = expr.col.name
    value_criteria = [ctx.col_range(value_col), ctx.cell(value_col)]
    matching = f"COUNTIFS({','.join(key_criteria + value_criteria)})"
    group_size = f"COUNTIFS({','.join(key_criteria)})"
    return f"{matching}={group_size}"


def compile_excel_value(expr: Expr, ctx: ExcelCompileContext) -> str:
    if isinstance(expr, ColExpr):
        return ctx.cell(expr.name)
    if isinstance(expr, IntLiteral):
        return str(expr.value)
    if isinstance(expr, StringLiteral):
        return _excel_string(expr.value)
    msg = f"unsupported value expression: {type(expr).__name__}"
    raise TypeError(msg)
