from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from validation_language_mockup.ast import Rule
from validation_language_mockup.evaluator import compile_rule, row_violation_mask
from validation_language_mockup.excel_compile import compile_violation_formula

_FILL_VIOLATION = PatternFill(fill_type="solid", fgColor="FFC7CE")


def default_excel_path(csv_file: Path) -> Path:
    return csv_file.with_name(f"{csv_file.stem}_validated.xlsx")


def export_validation_excel(
    df: pl.DataFrame,
    rule: Rule,
    path: Path,
) -> tuple[int, str]:
    """Write CSV data and apply row conditional formatting from the AVL rule."""
    formula = compile_violation_formula(rule, df.columns)
    compiled = compile_rule(rule)
    violation_count = sum(row_violation_mask(df, compiled))

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    ws.append(df.columns)
    for row_values in df.iter_rows():
        ws.append(list(row_values))

    if df.height > 0:
        last_col = get_column_letter(len(df.columns))
        data_range = f"A2:{last_col}{df.height + 1}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[formula], fill=_FILL_VIOLATION),
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return violation_count, formula
